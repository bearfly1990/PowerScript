#/usr/bin/python3
"""
author: xiche
create at: 04/09/2018
description:
    The progress to test the benchmark of import json messages
Change log:
Date        Author      Version     Description
04/09/2018  xiche       1.0         Set up this script
04/17/2018  xiche       1.0.1       Add the test result 'True/False' in 'TestResult.csv'
04/19/2018  xiche       1.0.2       Add function to get the date in the message and write to 'TestResult.csv'
05/22/2018  xiche       1.0.3       Separate config to different folder e.g. configs/PR1XPFIMORT01 configs/PR1XPFIAPP01
05/05/2018  xiche       1.0.4       Add CPU Memory Monitor
06/16/2018  xiche       1.0.5       Add Monitor class to do the monitor
                                    Change Test Result to shared folder.
06/21/2018  xiche       1.0.6       Add Import B pam messages
08/10/2018  xiche       1.0.7       Zip import message log file
                                    Add send email to users
                                    Monitor_Prolong
08/15/2018  xiche       1.0.8       Add and fix genereate test result and attach to email              
08/17/2018  xiche       1.0.9       Add client instance parameters
08/18/2018  xiche       1.1.0       Read restore db sql from Restore.sql
                                    Add MQ Setting to test result
                                    Add Client Num to test result
08/19/2018  xiche       1.1.1       Parameters could be comment in TestCases.ini
                                    Add Comment_Email settings in TestCases.ini
                                    Add cmutils_excel to style merged border
                                    Add running computer cpu cores and memroy size to test result.
"""
import sys
sys.path.insert(0, r"xxx\pythonlib")
import os
import glob
import configparser
import pymssql
import cx_Oracle
import time
import datetime
import csv
import shutil
import psutil
# import win32wnet
import subprocess
import re
import traceback
# import win32serviceutil
import queue as Queue
# import zipfile
from threading import Thread
import openpyxl
import xml.etree.ElementTree as ET
import copy

from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Border, Side, Font
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)


from cmutils.cmutils_value import isFloat
from cmutils.cmutils_log import Logger
from cmutils.cmutils_io import ConfigUtils, TxtUtils, PathUtils, CSVUtils, getDirFromFullPath
from cmutils.cmutils_monitor import Monitor
from cmutils.cmutils_db import DBUtils
from cmutils.cmutils_email import EmailUtils
from cmutils.cmutils_excel import style_range
# from lib.cmutils_monitor import CPUInfo, MemoryInfo
# import winreg
# import ctypes
# from ConfigParser import SafeConfigParser

# from subprocess import call
# config = configparser.ConfigParser()
# config = configParser.SafeConfigParser()
IMPORT_INI      = r"xxx\Import.ini"
COMPUTER_NAME   = os.environ['COMPUTERNAME']
COMPUTER_CPU_NUM = psutil.cpu_count()
COMPUTER_MEMORY_TOTAL = round(psutil.virtual_memory().total/1024/1024/1024)

IMPORT_INI      = IMPORT_INI.format(COMPUTER_NAME)

config          = configparser.RawConfigParser(strict=False)
config_db       = configparser.RawConfigParser(strict=False)
config_mq       = configparser.RawConfigParser(strict=False)
# RawConfigParser
DATE_TIME_FORMAT = "%Y/%m/%d %H:%M:%S" #"2018-06-13 23:51:17"
# config.read('%s\\%s\\Import.ini' % (CONFIG_SPACE,COMPUTER_NAME))
config.read(IMPORT_INI)

def timeout(second):
    os.system('timeout '+str(second))
    
def getTestResultLevel(x):
    return{
        'DIR':1,
        'FILE':2   
    }.get(x, 3)
    
def getCheckSql(x):
    return{
        '1':True,
        '0':False   
    }.get(x, False)
    
        
QUEUE_TEST_STATUS = Queue.Queue()
#QUEUE_TEST_STATUS.put(False)
'''
class Registry(object):
    def __init__(self, key_location, key_path):
        self.reg_key = winreg.OpenKey(key_location, key_path, 0, winreg.KEY_ALL_ACCESS)

    def set_key(self, name, value):
        try:
            _, reg_type = winreg.QueryValueEx(self.reg_key, name)
        except WindowsError:
            # If the value does not exists yet, we (guess) use a string as the
            # reg_type
            reg_type = winreg.REG_SZ
        winreg.SetValueEx(self.reg_key, name, 0, reg_type, value)

    def delete_key(self, name):
        try:
            winreg.DeleteValue(self.reg_key, name)
        except WindowsError:
            # Ignores if the key value doesn't exists
            pass
'''
TIME_IMPORT_START = datetime.datetime.now()
DEVMODE     = config['DevSetting']['Mode']
MULTI_VMS   = config['DevSetting']['MULTI_VMS']
RESTORE_DB  = config['DevSetting']['RESTORE_DB']
CLEAN_ENV   = config['DevSetting']['CLEAN_ENV']
SEND_EMAIL  = config['DevSetting']['SEND_EMAIL']

RUN_PAM_MESSAGE = config['DevSetting']['RUN_PAM_MESSAGE']
CMD_RUN_MANAGER = config['CMD']['StartServer']

# l:\pam\wsys\report\IMPORT.exe B C:\1.txt HZHAN T 5 42 Y 20170401 -i:m:\pam\wsys\pam.ini
CMD_RUN_IMPORT = config['CMD']['RunImport']
CMD_RUN_IMPORT_PAM = config['CMD']['RunImport_PAM']

if(RUN_PAM_MESSAGE == '1'):
    DEVMODE     = '1' 
    MULTI_VMS   = '0'
    RESTORE_DB  = '0' 
    CLEAN_ENV   = '0'
    CMD_RUN_IMPORT = CMD_RUN_IMPORT_PAM

DRIVER_L    = config['DriverMapping']['L']
DRIVER_M    = config['DriverMapping']['M']

# log
LOG_LEVEL   = config['LogConfig']['Level']
LOG_FORMAT  = config['LogConfig']['Format_Console']
LOG_FORMAT_FILE = config['LogConfig']['Format_File']
LOG_FILE    = config['LogConfig']['LogFile'].format(TIME_IMPORT_START)
LOG_DIR     = config['LogConfig']['LogDir']#.format(COMPUTER_NAME, TIME_IMPORT_START)

# Environment
KETTLE_HOME     = "KETTLE_HOME"
JAVA_HOME       = "JAVA_HOME"

EMAIL_RECIPIENTS    = config['EnvVar']['EMAIL_RECIPIENTS']

if(config.has_option('EnvVar', 'JAVA_HOME') and config['EnvVar']['JAVA_HOME'].strip()):
    os.environ["JAVA_HOME"]     = config['EnvVar']['JAVA_HOME']
JAVA_HOME                   = os.environ["JAVA_HOME"]

if(config.has_option('EnvVar', 'KETTLE_HOME') and config['EnvVar']['KETTLE_HOME'].strip()):
    os.environ["KETTLE_HOME"]   = config['EnvVar']['KETTLE_HOME'].format(COMPUTER_NAME)
KETTLE_HOME                 = os.environ["KETTLE_HOME"]
KETTLE_HOME                 = KETTLE_HOME.format(COMPUTER_NAME)

KETTLE_PROPERTY             = KETTLE_HOME + "/.kettle/kettle.properties"
KETTLE_PROPERTY_TPL         = KETTLE_PROPERTY + ".template"
#os.path.join
# reg = Registry(winreg.HKEY_LOCAL_MACHINE,r'SYSTEM\CurrentControlSet\Control\Session Manager\Environment')
# reg.set_key('KETTLE_HOME',KETTLE_HOME)

# os.putenv("JAVA_HOME", os.environ["JAVA_HOME"])
# os.putenv("KETTLE_HOME", os.environ["KETTLE_HOME"])
LOG4J_PATH                  = config['EnvVar']['LOG4J_PATH']
LOG4J_PATH                  = LOG4J_PATH.format(COMPUTER_NAME)
STTAPPMANAGER               = config['EnvVar']['STTAPPMANAGER']
MSG_TYPE_FLOW_MAPPING       = config['EnvVar']['MSG_TYPE_FLOW_MAPPING'].format(COMPUTER_NAME)
MSG_TYPE_FLOW_MAPPING_TPL   = MSG_TYPE_FLOW_MAPPING + ".template"

TEST_CASES_CONFIG_FILE      = config['EnvVar']['TEST_CASES_CONFIG'].format(COMPUTER_NAME)
TEST_CASE_LIST              = []

TRIGGER_LIST_WAIT           = config['EnvVar']['TRIGGER_LIST_WAIT']
TRIGGER_LIST_FINISHED       = config['EnvVar']['TRIGGER_LIST_FINISHED']
PAM_CORE                    = config['EnvVar']['PAM_CORE']
os.environ["PATH"]			= PAM_CORE + ";" + os.environ["PATH"]

PFS_ROOT                    = config['EnvVar']['PFS_ROOT']
PFS_ROOT                    = PFS_ROOT.format(COMPUTER_NAME)

MQ_CONFIG                   = config['EnvVar']['MQ_CONFIG']
mq_config_content = '[dummy_section]\n' + TxtUtils.read_string_from_txt(MQ_CONFIG)
config_mq.read_string(mq_config_content)
MQ_MAX_JAVA_MEMORY          = config_mq['dummy_section']['wrapper.java.maxmemory']

MSG_TYPE_FLOW_MAPPING_TO    = "{}\MsgTypeFlowMapping.txt".format(r'L:\PAM\Wsys\PDI_Transforms' if 'none' in PFS_ROOT.lower() else PFS_ROOT)#config['EnvVar']['MSG_TYPE_FLOW_MAPPING_TO']

DATA_SERVICE_CONFIG         = config['Files']['DATA_SERVICE_CONFIG'].format(COMPUTER_NAME)
POOLING_AGENT_CONFIG       = config['Files']['POOLING_AGENT_CONFIG'].format(COMPUTER_NAME)
DATA_SERVICE_CONFIG_L_PATH  = r"xxx\xxx.xml"
POOLING_AGENT_CONFIG_L_PATH = r"xxx\xxx.xml"
IMPORT_LIST                 = config['Files']['IMPORT_LIST'].format(COMPUTER_NAME)
TEST_RESULT                 = config['Files']['TEST_RESULT']#LOG_DIR + config['Files']['TEST_RESULT']
FILE_TEST_RESULT_TPL        = config['Files']['TEST_RESULT_TPL'].format(COMPUTER_NAME)
FILE_TEST_RESULT_XLSX       = TEST_RESULT + ".xlsx"
FILE_TEST_MONITOR           = config['Files']['TEST_MONITOR']#LOG_DIR + config['Files']['TEST_MONITOR']
IMPORT_RESULT_LOG           = config['Files']['IMPORT_RESULT_LOG']
IMPORT_RESULT_LOG_BK        = config['Files']['IMPORT_RESULT_LOG_BK']#LOG_DIR + config['Files']['IMPORT_RESULT_LOG_BK']
TEST_RESULT_LEVEL           = getTestResultLevel(config['TestResult']['Level'])
CHECK_SQL                   = getCheckSql(config['TestResult']['CheckSql'])
MONITOR_PROLONG             = config['TestResult']['Monitor_Prolong']
# KETTLE_SETTINS              = config['KettleSetting']
# EXPECTED_STT_MAX_ACTIVE_CONNECTIONS = config['KettleSetting']['STT_MAX_ACTIVE_CONNECTIONS']
KETTLE_SETTINGS             = []
for (each_key, each_val) in config.items('KettleSetting'):
    KETTLE_SETTINGS.append(each_val)
    
MESSAGE_TYPE_FLOW_MAPPINGS  = []
for (each_key, each_val) in config.items('MessageTypeFlowMapping'):
    MESSAGE_TYPE_FLOW_MAPPINGS.append(each_val)

'''Config logging'''
# logging.basicConfig(
    # level=LOG_LEVEL,
    # #format="%(asctime)s [%(threadName)-12.12s] [%(levelname)-5.5s]  %(message)s",
    # format=LOG_FORMAT,
    # handlers=[
        # logging.FileHandler("{0}/{1}".format(LOG_DIR,LOG_FILE)),
        # logging.StreamHandler()
# ])
format_file_str = "%(asctime)s [%(levelname)-5.5s] %(message)s"
format_console_str = "%(log_color)s%(asctime)s [%(levelname)-5.5s] %(message)s"
logfile_path_str = "{0}/{1}".format(LOG_DIR,LOG_FILE)
logging = None#Logger(level=LOG_LEVEL,format_file=format_file_str, format_console=format_console_str, logfile_path = logfile_path_str).logger

def checkDriverMapping():
    if(DEVMODE == '0'):
        subprocess.call("net use l: /delete /y", shell=True)
        subprocess.call("net use m: /delete /y", shell=True)
        os.system("net use l: {}".format(DRIVER_L))
        os.system("net use m: {}".format(DRIVER_M))
        print("Driver L: Deleted and Mapped:" +  DRIVER_L)
        print("Driver M: Deleted and Mapped:" +  DRIVER_M)
    else:
        try:
            pass
            # print("Driver L: Mapped:" +  win32wnet.WNetGetUniversalName("L:", 1))
            # print("Driver M: Mapped:" +  win32wnet.WNetGetUniversalName("M:", 1))
        except:
            print("Driver L: and M: Is Not Mapped")
            os.system("net use l: {}".format(DRIVER_L))
            os.system("net use m: {}".format(DRIVER_M))
checkDriverMapping()

FILE_RESTOE_DB_SQL  = config['DBInfo']['SQL_Restore_DB'].format(COMPUTER_NAME)

config_db.read(config['DBInfo']['Path'])
DB_INFO_CONFIG = config_db['DB']
# CONN = pymssql.connect(server='PFS-PFIDB-002\sql2012', 
               # user='METPDI', 
               # password='Pampfs2005', 
               # database='METPDI',
               # tds_version='8.0',
               # port = '1433')
DB_TYPE = "Sqlserver"

def getConnection():
    conn = None
    while(True):
        # try:             
            # CONN = cx_Oracle.connect('{}/{}@{}'.format(DB_INFO_CONFIG['userid'], DB_INFO_CONFIG['password'], config['DBInfo']['oracle_tnsname']))
            # DB_TYPE = "Oracle"
            # print("Oracle connected{}/{}@{}".format(DB_INFO_CONFIG['userid'], DB_INFO_CONFIG['password'], config['DBInfo']['oracle_tnsname']))
            # break
        # except Exception as e1:
        try:
        
            conn = pymssql.connect(DB_INFO_CONFIG['sqlnet'], DB_INFO_CONFIG['userid'], DB_INFO_CONFIG['password'], DB_INFO_CONFIG['source'])
            return conn
            break
        except Exception as e2:
            # print(traceback.format_exc())
            print("Connect to Database Error")
            print("try again...")
            timeout(10)    

    
def getDateInMessage(filePath): 
    dateStr = ""
    with open(filePath, "rt") as f:
        for line in f:
            searchObj = re.search("([0-9]{4}[0-9]{2}[0-9]{2})", line)
            if(searchObj):
                dateStr = searchObj.group(0)
            break
    return dateStr
def getLastNameFromFullPath(fullPath):
    lastIndex = fullPath.rfind('\\')
    if(lastIndex == -1):
        lastIndex = fullPath.rfind('/')
    return fullPath[lastIndex:]
'''c:\\aa\\bb\\cc\\file.txt => c:\\aa\\bb'''
def getLastDirPath(fullPath):
    lastIndex = fullPath.rfind('\\')
    lastIndex = fullPath.rfind('\\', 0, lastIndex - 1)
    if(lastIndex == -1):
        lastIndex = fullPath.rfind('/')
        lastIndex = fullPath.rfind('/', 0, lastIndex - 1)
    return fullPath[0:lastIndex]

def load_properties(filepath, sep='=', comment_char='#'):
    '''
    Read the file passed as parameter as a properties file.
    '''
    props = {}
    with open(filepath, "rt") as f:
        for line in f:
            l = line.strip()
            if l and not l.startswith(comment_char):
                key_value = l.split(sep)
                key = key_value[0].strip()
                value = sep.join(key_value[1:]).strip().strip('"') 
                props[key] = value 
    return props

def checkUser():
    # global CONN
    conn = getConnection()
    # CONN = pymssql.connect(DB_INFO_CONFIG['sqlnet'], DB_INFO_CONFIG['userid'], DB_INFO_CONFIG['password'], DB_INFO_CONFIG['source'])
    existFlag = False
    cursor = conn.cursor()
    # cursor.execute("SELECT dummy FROM sys.dual")
    # print(cursor.fetchone()[0])
    cursor.execute("select count(*) from pmuser where lower(keycode) = '{}'".format(os.getlogin()))
    # print("-----")
    # for row in cursor:
        # existFlag = True
        # print('row = %r' % (row,))
    # domain = os.environ['userdnsdomain']
    if cursor.fetchone()[0] >= 1:
        existFlag = True
    cursor.close()
    conn.close()
    if(existFlag):
        try:
            logging.info("DB:Server=%s,userid=%s,password=%s,database=%s",str(DB_INFO_CONFIG['sqlnet']), str(DB_INFO_CONFIG['userid']), str(DB_INFO_CONFIG['password']), str(DB_INFO_CONFIG['source']))
        except:
            logging.info("DB:Server=%s,userid=%s,password=%s,database=%s",str(DB_INFO_CONFIG['sqlnet']), str(DB_INFO_CONFIG['userid']), str(DB_INFO_CONFIG['password']), str(DB_INFO_CONFIG['schema']))
        logging.info("OK! - User {} is in the PAM DB".format(os.getlogin()))
    else:
        logging.error("Please confirm your login name is in the PAM DB!")
        exit()

def zipdir(path, ziph):
    # ziph is zipfile handle
    for root, dirs, files in os.walk(path):
        for file in files:
            ziph.write(os.path.join(root, file))

def checkImportResultLog(importFilePath):
    lins=[]
    recordsRead = -1
    loaded = -1
    processed = -1
    failedToLoad = -1
    failedToProcess = -1
    try:
        with open(IMPORT_RESULT_LOG) as f:
            lines = f.read().splitlines()
        for line in lines:
            line = line.strip()
            if("Records Read:" in line):
                recordsRead = line.split(":")[1].strip()
            if("Loaded:" in line):
                loaded = line.split(":")[1].strip()
            if("Processed:" in line):
                processed = line.split(":")[1].strip()
            if("Failed to Load:" in line):
                failedToLoad = line.split(":")[1].strip()
            if("Failed to Process:" in line):
                failedToProcess = line.split(":")[1].strip()
        logging.debug("Records Read:%s",recordsRead)  
        logging.debug("Loaded:%s",loaded)
        logging.debug("Processed:%s",processed)
        logging.debug("Failed to Load:%s",failedToLoad)
        logging.debug("Failed to Process:%s",failedToProcess)
        logging.info("Read:%s = Loaded:%s = Processed:%s | Failed Load=%s, Failed Process=%s",recordsRead,loaded,processed,failedToLoad,failedToProcess)
        if(int(failedToLoad) != 0):
            logging.error("Failed to Load:%s",failedToLoad)
        if(int(failedToProcess) != 0):
            logging.error("Failed to Process:%s",failedToProcess)
        if(int(loaded) <= 0 ):
            logging.error("No successed Loaded")
        if not os.path.exists(IMPORT_RESULT_LOG_BK):
            os.makedirs(IMPORT_RESULT_LOG_BK)
        moveToPath = IMPORT_RESULT_LOG_BK + getLastNameFromFullPath(importFilePath)+".ImportResult.log"
        if(os.path.exists(moveToPath)):
            shutil.move(IMPORT_RESULT_LOG, moveToPath+"_2.log")
        else:
            shutil.move(IMPORT_RESULT_LOG, moveToPath)

    except FileNotFoundError:
        logging.error("%s IS NOT EXISTED, PLEASE HAVE A CHECK", IMPORT_RESULT_LOG)
    return failedToProcess
    
def checkEnvVar(VarKey):
    try:
        logging.info(VarKey +"="+os.environ[VarKey])
    except KeyError:
        logging.error("THERE IS NO "+VarKey+" DEFINED")
    # finally:
        # print("finally")
        
def checkJAVAHOME():
    if(not JAVA_HOME.strip()):
        logging.error("Please set JAVA_HOME")
        exit()
    logging.info("JAVA_HOME={}".format(JAVA_HOME))
    # finally:
        # print("finally")
          
def checkKettleSetting():
    logging.info('#Kettle Settings:#')
    logging.info('KETTLE_HOME='+KETTLE_HOME)
    try:
        props   = load_properties(KETTLE_PROPERTY)
    except FileNotFoundError:
        logging.error("FILE <"+KETTLE_PROPERTY+"> NOT FOUND")
        exit()
    for key in KETTLE_SETTINGS:
        logging.info ('%s=' % key + props[key])

def checkPAMCore():
    _pam_core = PAM_CORE
    if(_pam_core.lower() in os.environ["PATH"].lower()):
        pass
    else:
        logging.error(PAM_CORE+" is not defined in PATH")
        exit()
    
def checkLog4j():
    logging.info(LOG4J_PATH)
    
def checkMsgTypeFlowMapping():
    logging.info("#MsgTypeFlowMapping Settings:#")
    if(not "none" in PFS_ROOT.lower()):
        M_PAM_INI             = config['EnvVar']['M_PAM_INI']
        ConfigUtils.set(M_PAM_INI, "Securities", "PFS_ROOT", PFS_ROOT)
    with open(MSG_TYPE_FLOW_MAPPING_TO, "rt") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#'):
                for msg in MESSAGE_TYPE_FLOW_MAPPINGS:
                    if( msg in line):
                        logging.info(line)
    # config_mpamini      = configparser.RawConfigParser(strict=False, allow_no_value=True)
    # config_mpamini.read(MPAMINI)
    # config_mpamini.set('Securities', 'PFS_ROOT', PFS_ROOT)
    # with open(MPAMINI, 'w') as configfile:
        # config_mpamini.write(configfile)
    # try:
        # shutil.copy(MSG_TYPE_FLOW_MAPPING, MSG_TYPE_FLOW_MAPPING_TO)
    # except PermissionError:
        # logging.info("%s IS IN USED!", MSG_TYPE_FLOW_MAPPING_TO)
        # goon = input("CONTINUE?(Y/N): ")
        # if(goon.upper() == "Y"):
            # pass
        # else:
            # logging.info("PLEASE CHANGE THE SETTINGS FIRST!")
            # exit()   
    # try:
        # pass
        # shutil.copy(MSG_TYPE_FLOW_MAPPING, MSG_TYPE_FLOW_MAPPING_TO)
        
        # with open(MSG_TYPE_FLOW_MAPPING_TO, "rt") as f:
            # for line in f:
                # l = line.strip()
                # if l and not l.startswith('#'):
                    # if("TABLOCKREGULARLOANPAYMENT;" in l):
                        # logging.info(l)
    # except Exception as e:
        # logging.error(traceback.format_exc())
    # except PermissionError:
        # logging.error("No Permission to replace the %s", MSG_TYPE_FLOW_MAPPING_TO)
        # exit()
   

def readConfigFile(filePath='ImportConfig.txt'):
    with open(filePath) as f:
        content = f.read().splitlines();
    return [x.strip() for x in content if (not "#" in x and x.strip())]
    
def countInvalidRowsInFile(filePath):
    count = 0
    try:
        with open(filePath) as f:
            content = f.read().splitlines()
        for line in content:
            if(line.strip()):
                count = count + 1
    except FileNotFoundError:
        logging.error("FILE %s NOT FOUND!", filePath)
    return count
    
def initJavaServer():
    logging.info("start init Java Server...")
    timeout(5)
    _cmd_run_manager = ""
    if(DEVMODE == '0'):
        _cmd_run_manager = CMD_RUN_MANAGER.format(JAVA_HOME, LOG4J_PATH, KETTLE_HOME, STTAPPMANAGER)
        logging.info("CMD_RUN_MANAGER="+_cmd_run_manager)
        os.system(_cmd_run_manager)
        timeout(480)
    
def getSqlStr(dir):
    checkSqlFile1 = dir+"\\checkResult.sql"
    checkSqlFile2 = getLastDirPath(checkSqlFile1)+"\\checkResult.sql"
    checkSqlFile3 = getLastDirPath(checkSqlFile2)+"\\checkResult.sql"
    sqlStr = "select -1"
    try:
        with open(checkSqlFile1) as f:
            # content = f.read().splitlines()
            sqlStr  = f.read().replace('\n', '')
    except FileNotFoundError:
        try:
            with open(checkSqlFile2) as f:
                # content = f.read().splitlines();
                # sqlStr = content[0]
                sqlStr  = f.read().replace('\n', '')
        except FileNotFoundError:
            try:
                with open(checkSqlFile3) as f:
                    # content = f.read().splitlines();
                    # sqlStr = content[0]
                    sqlStr  = f.read().replace('\n', '')
            except FileNotFoundError:
                logging.error("File %s is not found!", checkSqlFile1)
                logging.error("File %s is not found!", checkSqlFile2)
                logging.error("File %s is not found!", checkSqlFile3)
    return sqlStr
    
def getCurrentRowsInDB(dir):
    if CHECK_SQL:
        pass
    else:
        return 0
    resultRows = 0
    cursor = CONN.cursor()
    sqlStr = getSqlStr(dir)
    try:
        cursor.execute(sqlStr)
        for row in cursor:
            resultRows = row[0]
            break
    except Exception:
        logging.error("SQL ERROR:%s", sqlStr)
    return resultRows
    
def checkImportResultInDB(dir, rowsBefore, rowsAfter, expectedRows):
    if CHECK_SQL:
        pass
    else:
        return 0
    resultRows = rowsAfter - rowsBefore
    rowsAreMatched = (resultRows == expectedRows)
    logging.info("Rows DB Added:%s = %s - %s", resultRows, rowsAfter, rowsBefore)
    logging.info("Rows Expected:%s", expectedRows)
    
    if(not rowsAreMatched or expectedRows == 0):
        logging.error("Rows not matched!")
        logging.error("Failed Import %s", dir)
    else:
        logging.info("Successed Import %s", dir)
    return rowsAreMatched
        
def writeResult(pathStr, rows, time, messageStr, isDIR, rowsAreMatched_DB, rowsAreMatched_LOG, failedToProcess):
    writeFlag = False
    TEST_RESULT_LEVEL == 1 # DIR = 1 FILE = 2
    # if (isDIR and ".txt" in pathStr):
        # writeFlag = False
    if(TEST_RESULT_LEVEL == 1 and isDIR):
        writeFlag = True
    if(TEST_RESULT_LEVEL == 2):
        writeFlag = True
        
    if(writeFlag):
        timeStr = "{:.2f}".format(time)
        timeStrMinutes = "{:.2f}".format(time/60)
        timAvgStr = "-1"
        if(rows > 0):
            timAvgStr = "{:.4f}".format(time/rows)
        '''    
            with open(TEST_RESULT, 'a') as writer:
                writer.write("%s :\n\ttotal:%s avg:%s rows:%s\n" % (pathStr,timeStr, timAvgStr, rows))
        '''
        myData = [pathStr,rows, timeStr, timeStrMinutes,timAvgStr, messageStr, rowsAreMatched_DB,rowsAreMatched_LOG, failedToProcess]
        myFile = open(TEST_RESULT, 'a', newline='')  
        with myFile:  
            writer = csv.writer(myFile)
            writer.writerow(myData)    
        
def initTestResult():
    myData = ['path', 'rows', 'timeused(s)','timeused(m)','avgtime', 'messageDate', 'rowsAreMatchedInDB', 'rowsAreMatchedByLog', 'failedToProcess']
    # csv.register_dialect('myDialect', delimiter=',', quoting=csv.QUOTE_NONE)
    # myFile = open(TEST_RESULT, 'w')  
    # with myFile:  
        # writer = csv.writer(myFile, dialect='myDialect')
   
    myFile = open(TEST_RESULT, 'w', newline='')  
    with myFile:  
        writer = csv.writer(myFile)
        writer.writerow(myData)

def zip_import_logs():
    if(os.path.exists(IMPORT_RESULT_LOG_BK)):
        shutil.make_archive(IMPORT_RESULT_LOG_BK, 'zip', IMPORT_RESULT_LOG_BK)
        # zipf = zipfile.ZipFile(IMPORT_RESULT_LOG_BK+'.zip', 'w', zipfile.ZIP_DEFLATED)
        # zipdir(IMPORT_RESULT_LOG_BK, zipf)
        # zipf.close()
        shutil.rmtree(IMPORT_RESULT_LOG_BK)
    
def runImport():
    initTestResult()
    _cmd_run_import = ""
    dirs = readConfigFile(IMPORT_LIST)
    list_of_files = []
    for dir in dirs:
        logging.info("-----START IMPORT-----")
        countRowsDir = 0
        list_of_files = glob.glob(dir+'/*.txt') 
        list_of_files.sort()
        rowsInDBBefore_Dir = getCurrentRowsInDB(dir)
        timeStartDir  =  time.time()
        dirTimeBack = 0
        messageDateStr = ''
        failedToProcessDir = 0
        if(len(list_of_files) > 0):
            messageDateStr = getDateInMessage(list_of_files[0])
        rowsAreMatched = False
        for importfile in list_of_files:
            rowsInFile = countInvalidRowsInFile(importfile)
            rowsInDBBefore_File = getCurrentRowsInDB(dir)
            countRowsDir += rowsInFile
            countRowsFile = rowsInFile
            timeStartFile  =  time.time()
            
            _cmd_run_import = CMD_RUN_IMPORT.format(importfile)
            logging.info(_cmd_run_import)
            #Real to call the import
            if(DEVMODE == '0' or DEVMODE == '1'):
                os.system(_cmd_run_import)
            timeEndFile  =  time.time()
            
            periodFile  = timeEndFile - timeStartFile
            periodFileStr = "{:.2f}".format(periodFile)
            logging.info("Finished:"+importfile)
            logging.info("Import Time(s):%s", periodFileStr)
            
            timeout(10)
            dirTimeBack = dirTimeBack + 10
            
            rowsInDBAfter_File = getCurrentRowsInDB(dir)
            rowsAreMatched = checkImportResultInDB(importfile, rowsInDBBefore_File, rowsInDBAfter_File, rowsInFile)
            # time.sleep(5)
            '''Change the time back because sleep'''
            failedToProcessFile = int(checkImportResultLog(importfile))
            failedToProcessDir  = failedToProcessDir + failedToProcessFile
            writeResult(importfile, rowsInFile, periodFile, messageDateStr, False, rowsAreMatched,failedToProcessFile == 0, failedToProcessFile)
            
        timeEndDir  =  time.time()
        #{:10.4f}
        periodDir = timeEndDir - timeStartDir - dirTimeBack
        periodDirStr = "{:.2f}".format(periodDir)
        logging.info("Finished:"+dir)
        logging.info("Import Time(s):%s", periodDirStr)
        
        rowsInDBAfter_Dir = getCurrentRowsInDB(dir)
        
        rowsAreMatched = checkImportResultInDB(dir, rowsInDBBefore_Dir, rowsInDBAfter_Dir, countRowsDir)
        
        logging.info("-----END IMPORT-----")
        writeResult(dir, countRowsDir, periodDir, messageDateStr, True, rowsAreMatched, failedToProcessDir == 0, failedToProcessDir)
        
        if(DEVMODE == '0' or DEVMODE == '1'):
            timeout(5)
            dirTimeBack = dirTimeBack - 5
    zip_import_logs()
            
def startMonitor():
    Monitor(FILE_TEST_MONITOR, QUEUE_TEST_STATUS).startMonitor()
    
def readTestCasesConfig():
    global TEST_CASE_LIST
    testcase_config = configparser.RawConfigParser(strict=False)

    testcase_config.read(TEST_CASES_CONFIG_FILE)
    TEST_CASE_LIST = []
    
    
    
    
    for section in testcase_config.sections():
        testcase = type('',(object,),{\
        'pool_size':testcase_config.get(section, "PoolSize", fallback=""),\
        'pool_size2':testcase_config.get(section, "PoolSize2", fallback=""),\
        'algo_pool_size':testcase_config.get(section, "STT_ALGO_POOL_SIZE", fallback=""),\
        'java_memory':testcase_config.get(section, "Java_Memory", fallback=""),\
        'PARAM1':testcase_config.get(section, "PARAM1", fallback=""),\
        'PARAM2':testcase_config.get(section, "PARAM2", fallback=""),\
        'PARAM3':testcase_config.get(section, "PARAM3", fallback=""),\
        'client_num':testcase_config.get(section, "ClientNum", fallback=""),\
        'email_comment':testcase_config.get(section, "Email_Comment", fallback="")\
        })()
         # testcase_config[section]["PARAM1"],\
         # testcase_config[section]["PARAM2"],\
         # testcase_config[section]["PARAM3"],\
        TEST_CASE_LIST.append(testcase)
        
def generateConfigFiles(test_case):
    TxtUtils.replaceVariables(MSG_TYPE_FLOW_MAPPING_TPL, MSG_TYPE_FLOW_MAPPING, {"%POOL_SIZE%": test_case.pool_size, "%POOL_SIZE2%": test_case.pool_size2})
    TxtUtils.replaceVariables(KETTLE_PROPERTY_TPL, KETTLE_PROPERTY, {"%PC_NAME%": COMPUTER_NAME, "%STT_ALGO_POOL_SIZE%":test_case.algo_pool_size})
    update_client_instance(int(test_case.client_num))
    shutil.copy(MSG_TYPE_FLOW_MAPPING, os.path.join(LOG_DIR, "MsgTypeFlowMapping.txt"))
    shutil.copy(KETTLE_PROPERTY, os.path.join(LOG_DIR, "kettle.properties"))

def update_client_instance(client_num = 1):
    # et = xml.etree.ElementTree.parse(data_service_config)
    logging.info("Try to change client instance to: {}".format(client_num))
    data_service_et     = ET.parse(DATA_SERVICE_CONFIG)
    pooling_agent_et    = ET.parse(POOLING_AGENT_CONFIG)

    root_maps           = data_service_et.getroot()
    root_configurations = pooling_agent_et.getroot()

    el_map = root_maps.find(".//Map")
    for i in range(client_num - 1):
        temp_el_map         = copy.deepcopy(el_map)
        temp_el_clientID    = temp_el_map.find(".//ClientID")
        temp_el_clientID.text =  temp_el_clientID.text + str(i+1)
        root_maps.append(temp_el_map)


    el_client_specific_configuration = root_configurations.find(".//ClientSpecificConfiguration")
    el_client = root_configurations.find(".//ClientSpecificConfiguration/Client")

    for i in range( client_num - 1 ):
        temp_el_client  = copy.deepcopy(el_client)
        temp_clientID   = temp_el_client.get("ClientID")
        temp_el_client.set("ClientID", temp_clientID + str(i+1))
        el_client_specific_configuration.append(temp_el_client)
    try:
        data_service_et.write(DATA_SERVICE_CONFIG_L_PATH)
        pooling_agent_et.write(POOLING_AGENT_CONFIG_L_PATH)
    except PermissionError:
        logging.warn("No permission to update File {}".format(DATA_SERVICE_CONFIG_L_PATH))
        logging.warn("No permission to update File {}".format(POOLING_AGENT_CONFIG_L_PATH))
        logging.warn("Client Instance Number Could Not be Changed!")
        
def update_configs(test_case):
    global CMD_RUN_MANAGER
    global TIME_IMPORT_START
    global LOG_DIR
    # global IMPORT_LIST
    global TEST_RESULT
    global FILE_TEST_RESULT_XLSX
    global FILE_TEST_MONITOR
    global IMPORT_RESULT_LOG_BK
    global logging
    global logfile_path_str
    
    CMD_RUN_MANAGER         = CMD_RUN_MANAGER.replace("%JAVA_MEMORY%", test_case.java_memory)
    CMD_RUN_MANAGER         = CMD_RUN_MANAGER.replace("%PARAM1%", test_case.PARAM1)
    CMD_RUN_MANAGER         = CMD_RUN_MANAGER.replace("%PARAM2%", test_case.PARAM2)
    CMD_RUN_MANAGER         = CMD_RUN_MANAGER.replace("%PARAM3%", test_case.PARAM3)
    TIME_IMPORT_START       = datetime.datetime.now()
    LOG_DIR                 = config['LogConfig']['LogDir'].format(COMPUTER_NAME, TIME_IMPORT_START) + "_{}_{}_{}".format(test_case.pool_size, test_case.algo_pool_size, test_case.java_memory)
    # IMPORT_LIST             = os.path.join(LOG_DIR, config['Files']['IMPORT_LIST'].format(COMPUTER_NAME))
    TEST_RESULT             = os.path.join(LOG_DIR, TEST_RESULT)
    FILE_TEST_RESULT_XLSX   = os.path.join(LOG_DIR, FILE_TEST_RESULT_XLSX)
    FILE_TEST_MONITOR       = os.path.join(LOG_DIR, FILE_TEST_MONITOR)
    IMPORT_RESULT_LOG_BK    = os.path.join(LOG_DIR, IMPORT_RESULT_LOG_BK)
    logfile_path_str        = os.path.join(LOG_DIR, LOG_FILE)
    
    if not os.path.exists(LOG_DIR):
        os.makedirs(LOG_DIR)
        
    logging = Logger(level=LOG_LEVEL,format_file=format_file_str, format_console=format_console_str, logfile_path = logfile_path_str).logger

    generateConfigFiles(test_case)
    
def restore_db():
    if(DEVMODE != '0' or RESTORE_DB == '0'):
        return 0
    logging.info("start to restore db...")
    timeout(5)
    db_utils = DBUtils(DB_INFO_CONFIG['sqlnet'], DB_INFO_CONFIG['userid'], DB_INFO_CONFIG['password'], DB_INFO_CONFIG['source'])
    logging.info("Start to restore DB {}/{}".format(DB_INFO_CONFIG['sqlnet'], DB_INFO_CONFIG['source']))
    # os.system("schtasks.exe /Run /S sstcp18264-7.nj.pfs.net /TN restoreDB")
    # os.system("start python test.py")
    sql_restore_db = TxtUtils.read_string_from_txt(FILE_RESTOE_DB_SQL)
    try:
        db_utils.restore_db(sql_restore_db)
    except:
        logging.warn("DB Exception! wait 10 minutes")
        timeout(600)

def cleanEnvironment():
    if(DEVMODE != '0' or CLEAN_ENV == '0'):
        return 0
    os.system("tskill AlgoQueueService /a")
    os.system("tskill java /a")
    os.system("tskill import /a")
    service = None
    try:
        service = psutil.win_service_get("ActiveMQ")
        service = service.as_dict()
        logging.info("Restart Service ActiveMQ...")
    except:
        try:
            service = psutil.win_service_get("ActiveMQService")
            service = service.as_dict()
            logging.info("Restart Service ActiveMQService...")
        except:
            logging.warn("Service ActiveMQ or ActiveMQService not exist")
            return 0
    os.system("net stop {}".format(service['name']))
    os.system("net start {}".format(service['name']))
    # win32serviceutil.RestartService(serviceName)
    
def vm_trigger(status):
    if(MULTI_VMS == "0"):
        return 0
    if(status == "start"):
        print("check and wait trigger...")
        while(True):
            vm_wait = TxtUtils.read_first_line(TRIGGER_LIST_WAIT)
            print("vm_wait" + "-->" + vm_wait)
            if(COMPUTER_NAME.lower() in vm_wait.lower()):
                break
            else:
                # print("Wait for last vm finished in 10 seconds...")
                timeout(10)
    else:
        TxtUtils.write_list_to_file_with_newline(TRIGGER_LIST_FINISHED, ["{} {}".format(TIME_IMPORT_START,COMPUTER_NAME)], mode='a')
        TxtUtils.remove_first_line(TRIGGER_LIST_WAIT)
def sendTestResult(test_case):
    if(SEND_EMAIL == "0"):
        return 0
    logging.info("send test result...")
    emailUtil = EmailUtils()
    emailUtil.set_recipients(EMAIL_RECIPIENTS)
    emailUtil.add_attachment(FILE_TEST_RESULT_XLSX)
    emailUtil.set_body("<p>{}</p><p><a href='{}'>{}</a></p>".format(test_case.email_comment, LOG_DIR,LOG_DIR))
    emailUtil.send_email()
    
def checkEnvAndSettings():
    checkUser()
    checkPAMCore()
    checkJAVAHOME()
    checkKettleSetting()
    checkLog4j()
    checkMsgTypeFlowMapping()
    # goon = input("ARE THE SETTINGS OK?(Y/N): ")
    # if(goon.upper() == "Y"):
        # pass
    # else:
        # logging.warn("PLEASE CHANGE THE SETTINGS FIRST!")
        # QUEUE_TEST_STATUS.put(True)
        # exit()
def generate_test_result(test_case):
    global FILE_TEST_RESULT_XLSX
    file_monitor            = FILE_TEST_MONITOR #os.path.join(LOG_DIR, "Monitor.csv")
    file_test_result        = TEST_RESULT #os.path.join(LOG_DIR, "TestResult.csv")
    os_memory_max = 0
    os_memory_last = 0
    java_memory_last = 0
    logging.info("Generate ==> {}".format(FILE_TEST_RESULT_XLSX))
    # wb_tpl = Workbook()
    wb_tpl  = openpyxl.load_workbook(FILE_TEST_RESULT_TPL)
    ws      = wb_tpl['Data'] 
    """ [ [2018/06/17 23:47:11,17.0,16.71], [2018/06/17 23:47:13,1.5,17.01], [2018/06/17 23:47:15,2.9,17.04], ... ] """
    data_list = CSVUtils.readCSVRowsList(file_monitor)
    _max_row = len(data_list)
    _max_col = 5
    for row_index, rows in enumerate(data_list):
        for col_index, value in enumerate(rows):
            if(row_index == 0):
                ws.cell(row=row_index+1, column=col_index+1).value = value
            else:
                if(col_index == 0):
                    ws.cell(row=row_index+1, column=col_index+1).value = datetime.datetime.strptime(value, DATE_TIME_FORMAT)
                    ws.cell(row=row_index+1, column=col_index+1).number_format = 'HH:mm:ss'
                else:
                    ws.cell(row=row_index+1, column=col_index+1).value = float(value)
                    if(col_index == 2):#os memory
                        os_memory_max   = max(float(value), os_memory_max)
                        os_memory_last  = float(value)
                    if(col_index == 3):
                        java_memory_last = float(value)
            
    # ws['E1'] = "CPU(AVG):"
    # ws['E2'] = "Memory(AVG):"
    # ws['F1'] = "=AVERAGE(B:B)"
    # ws['F2'] = "=AVERAGE(C:C)"
    chart1 = LineChart()
    chart1.title = "CPU/Memory Monitor"
    chart1.style = 2 # default style when new a line chart
    chart1.height = 10 # default is 7.5
    chart1.width = 35 # default is 15
    chart1.legend.position = "b"
    chart1.y_axis.scaling.min = 0
    chart1.y_axis.scaling.max = 100


    # set y-axis
    data = Reference(ws, min_col=2, min_row=1, max_col=_max_col, max_row=_max_row)
    chart1.add_data(data, titles_from_data=True)

    # set time as category(x-axis)
    cats = Reference(ws, min_col=1, min_row=2, max_row=_max_row)
    chart1.set_categories(cats)    

    s2 = chart1.series[0]
    s2.smooth = True # Make the line smooth

    s2 = chart1.series[1]
    s2.smooth = True # Make the line smooth

    ws_test_result = wb_tpl['TestResult']

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws_test_result['B1'] = os_memory_max*16/100
    ws_test_result['B2'] = os_memory_last*16/100
    ws_test_result['B3'] = java_memory_last*16/100

    ws_test_result['A4'] = "{}Cores {}GB {}({}:{})".format(COMPUTER_CPU_NUM, COMPUTER_MEMORY_TOTAL, COMPUTER_NAME, DB_INFO_CONFIG['sqlnet'], DB_INFO_CONFIG['source'])
    
    style_range(ws_test_result, 'A4:F4', border = border)

    ws_test_result['A6'] = test_case.pool_size
    ws_test_result['B6'] = test_case.algo_pool_size
    ws_test_result['C6'] = MQ_MAX_JAVA_MEMORY + "MB"
    ws_test_result['D6'] = test_case.java_memory
    ws_test_result['E1'] = test_case.client_num
    ws_test_result['E6'] = "{} {} {}".format(test_case.PARAM1, test_case.PARAM2, test_case.PARAM3)
    
    
    ws_test_result['E5'].border = border
    ws_test_result['F5'].border = border
    ws_test_result['E6'].border = border
    ws_test_result['F6'].border = border
    data_list = CSVUtils.readCSVRowsList(file_test_result)

    row_start = 8
    col_start = 1
    row_count = 0
    col_count = 0

    import_num = len(data_list)
    cols_total = 0
    
    ws_test_result.add_chart(chart1, "A{}".format(import_num  + row_start ))

    for row_index, rows in enumerate(data_list):
        if(row_index == 0):
            continue
        cols_total = len(rows)
        
        col_count = 0
        for col_index, value in enumerate(rows):
            if(col_index in [5, 6, 7]):
                continue
            if(isFloat(value)):
                ws_test_result.cell(row=row_count + row_start, column=col_count + col_start).value = float(value)
            elif(col_index == 0):
                ws_test_result.cell(row=row_count + row_start, column=col_count + col_start).value = PathUtils.getFileNameFromFullPath(value)
            else:
                ws_test_result.cell(row=row_count + row_start, column=col_count + col_start).value = value
            ws_test_result.cell(row=row_count + row_start, column=col_count + col_start).border = border   
            col_count = col_count + 1
        row_count = row_count + 1

    cols_total = cols_total - len( [5, 6, 7])

    for sum_col_index in range(cols_total):
        ws_test_result.cell(row = row_count + row_start, column = sum_col_index+1).border = border
        ws_test_result.cell(row = row_count + row_start, column = sum_col_index+1).font = Font(bold=True)

    ws_test_result.cell(row = row_count + row_start, column = 1).value = "sum"

    ws_test_result.cell(row = row_count + row_start, column = 4).value = "=sum(D{}:D{})".format(row_start, row_start + row_count-1)
    
    wb_tpl.save(filename = FILE_TEST_RESULT_XLSX)
    
def __main__():
    # checkDriverMapping() have move ahead at line 184
    vm_trigger("start")
    print("Read test cases config...")
    readTestCasesConfig()
    for test_case in TEST_CASE_LIST:
        print("Update configs...")
        update_configs(test_case)
        restore_db()
        startMonitor()
        logging.info("-----START TO RUN IMPORT PROCESS!-----")
        logging.info("-----CHECK SETTINGS-----")
        checkEnvAndSettings()
        initJavaServer()
        runImport()
        print("Import Finished!")
        timeout(int(MONITOR_PROLONG))
        QUEUE_TEST_STATUS.put(True)
        generate_test_result(test_case)
        sendTestResult(test_case)
        cleanEnvironment()
    vm_trigger("ended")
__main__()
