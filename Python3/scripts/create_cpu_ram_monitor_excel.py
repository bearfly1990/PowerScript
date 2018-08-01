import sys
# sys.path.insert(0, "\\\\nj.pfs.net\\departments\\Development\\Team Engineering\\xiche\\pythonlib")
sys.path.insert(0, r"\\nj.pfs.net\departments\Development\Team Engineering\xiche\pythonlib")
import openpyxl
from datetime import datetime
from cmutils.cmutils_io import CSVUtils
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis
from openpyxl import Workbook
import glob
import os



DATE_TIME_FORMAT = "%Y/%m/%d %H:%M:%S" #"2018-06-13 23:51:17"
    
def test_monitor(file_in, file_out):
    print("Generate {} ==> {}".format(file_in, file_out))
    wb_tpl = Workbook()
    ws = wb_tpl.active
    # ws = wb_tpl.worksheets[0]
    ws.title = "Monitor"
    """ [ [2018/06/17 23:47:11,17.0,16.71], [2018/06/17 23:47:13,1.5,17.01], [2018/06/17 23:47:15,2.9,17.04], ... ] """
    data_list = CSVUtils.readCSVRowsList(file_in)
    _max_row = len(data_list)
    _max_col = 3
    for row_index, rows in enumerate(data_list):
        for col_index, value in enumerate(rows):
            if(row_index == 0):
                ws.cell(row=row_index+1, column=col_index+1).value = value
            else:
                if(col_index == 0):
                    ws.cell(row=row_index+1, column=col_index+1).value = datetime.strptime(value, DATE_TIME_FORMAT)
                    ws.cell(row=row_index+1, column=col_index+1).number_format = 'HH:mm:ss'
                else:
                    ws.cell(row=row_index+1, column=col_index+1).value = float(value)
                    
    ws['E1'] = "CPU(AVG):"
    ws['E2'] = "Memory(AVG):"
    ws['F1'] = "=AVERAGE(B:B)"
    ws['F2'] = "=AVERAGE(C:C)"
    chart1 = LineChart()
    chart1.title = "CPU/Memory Monitor"
    chart1.style = 2 # default style when new a line chart
    chart1.height = 15 # default is 7.5
    chart1.width = 30 # default is 15
    chart1.legend.position = "b"
    chart1.y_axis.scaling.min = 0
    chart1.y_axis.scaling.max = 100
    # chart1.y_axis.title = 'Pecent'
    # chart1.x_axis.title = 'Time'

    # set y-axis
    data = Reference(ws, min_col=2, min_row=1, max_col=_max_col, max_row=_max_row)
    chart1.add_data(data, titles_from_data=True)

    # set time as category(x-axis)
    cats = Reference(ws, min_col=1, min_row=2, max_row=_max_row)
    chart1.set_categories(cats)    

    # Style the lines
    # s1 = c1.series[0]
    # s1.marker.symbol = "triangle"
    # s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
    # s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline

    # s1.graphicalProperties.line.noFill = True

    s2 = chart1.series[0]
    s2.smooth = True # Make the line smooth
    # s2.graphicalProperties.line.solidFill = "00AAAA"
    # s2.graphicalProperties.line.dashStyle = "sysDot"
    # s2.graphicalProperties.line.width = 100050 # width in EMUs

    s2 = chart1.series[1]
    s2.smooth = True # Make the line smooth
    # chart1.x_axis.tickLblPos = "low"
    # chart1.x_axis.tickLblSkip = 3 # whatever you like
    ws.add_chart(chart1, "F10")
   
            
    # dest_filename = file_out
    wb_tpl.save(filename = file_out)
    print("%s ====> %s" % (file_in, file_out))
    
STR_FIND = "Shawn"
CASE_SENSITIVE = True
FILE_PATTERN = '**\Monitor.csv'

list_of_files = glob.glob(FILE_PATTERN,recursive=True)

for file_name in list_of_files:
    if(os.path.isfile(file_name+".xlsx")):
        continue
    test_monitor(file_name, file_name+".xlsx")

    
